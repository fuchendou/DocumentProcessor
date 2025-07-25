import os
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from core import DocumentProcessor
from exceptions import FileCorruptionError
from utils.file_utils import validate_file_exists

class XLSXProcessor(DocumentProcessor):
    """处理Excel格式(XLSX)的处理器"""
    SUPPORTED_EXTENSIONS = ['xlsx', 'xlsm']
    
    def __init__(self, file_path: str, unique_key: str, max_len: int):
        super().__init__(file_path)
        validate_file_exists(file_path)
        self.file_extension = os.path.splitext(file_path)[1][1:].lower()
        self.unique_key = unique_key
        self.max_len = max_len
        self.workbook = None
        self.merged_cells_cache = {}  # 缓存合并单元格信息
    
    def extract_text(self) -> str:
        """提取所有工作表的文本内容，用换页符分隔"""
        if self.file_extension in self.SUPPORTED_EXTENSIONS:
            return self._extract_xlsx_text()
    
    def extract_metadata(self) -> dict:
        """提取Excel元数据"""
        try:
            self._load_workbook()
            meta = {
                'file_type': self.file_extension,
                'sheet_count': len(self.workbook.sheetnames),
                'sheet_names': self.workbook.sheetnames,
                'has_formulas': self._check_has_formulas(),
                'has_hyperlinks': self._check_has_hyperlinks(),
                'has_merged_cells': self._check_has_merged_cells()
            }
            return meta
        except Exception as e:
            raise FileCorruptionError(f"Error reading XLSX metadata: {e}") from e
    
    def _load_workbook(self):
        """加载工作簿（延迟加载）"""
        if not self.workbook:
            try:
                self.workbook = openpyxl.load_workbook(
                    self.file_path, 
                    read_only=True,   # 只读模式提高性能
                    data_only=True,   # 获取公式计算值
                    keep_links=False  # 不保留外部链接
                )
            except Exception as e:
                raise FileCorruptionError(f"Invalid XLSX file: {e}") from e
    
    def _extract_xlsx_text(self) -> str:
        """提取XLSX文本内容"""
        try:
            self._load_workbook()
            content = []
            
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                # 预加载合并单元格信息
                self._load_merged_cells(sheet_name, sheet)
                sheet_content = self._process_sheet(sheet_name, sheet)
                if sheet_content:
                    content.append(f"=== Sheet: {sheet_name} ===")
                    content.append(sheet_content)
            
            return '\n'.join(content)
        except Exception as e:
            raise FileCorruptionError(f"Error processing XLSX: {e}") from e
        finally:
            # 确保关闭工作簿释放资源
            if self.workbook:
                self.workbook.close()
                self.workbook = None
            self.merged_cells_cache.clear()
    
    def _load_merged_cells(self, sheet_name: str, sheet):
        """加载并缓存合并单元格信息（兼容只读模式）"""
        if sheet_name in self.merged_cells_cache:
            return
            
        merged_map = {}
        # 使用merged_cell_ranges代替merged_cells属性
        for merged_range in sheet.merged_cells.ranges if hasattr(sheet, 'merged_cells') else []:
            self._add_merged_range(merged_map, merged_range)
        
        # 处理只读模式下的合并单元格
        if hasattr(sheet, 'merged_cell_ranges'):
            for range_str in sheet.merged_cell_ranges:
                cr = CellRange(range_str)
                self._add_merged_range(merged_map, cr)
        
        self.merged_cells_cache[sheet_name] = merged_map
    
    def _add_merged_range(self, merged_map: dict, merged_range):
        """添加合并区域到映射表"""
        top_left = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != top_left:
                    merged_map[(row, col)] = top_left
    
    def _process_sheet(self, sheet_name: str, sheet) -> str:
        """处理单个工作表"""
        MAX_LEN = self.max_len
        UNIQUE_KEY = self.unique_key
        OVERHEAD_PADDING = 3
        content = []
        
        # 获取表头（第一行）
        headers = self._get_headers(sheet)
        
        # 查找唯一标识列索引
        unique_index = 0
        if UNIQUE_KEY and headers:
            try:
                unique_index = headers.index(UNIQUE_KEY)
            except ValueError:
                unique_index = 0
        
        # 获取合并单元格映射
        merged_map = self.merged_cells_cache.get(sheet_name, {})
        
        # 处理每一行
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # 跳过空行
            if all(cell is None or cell == '' for cell in row):
                continue
                
            # 获取唯一标识
            unique_id = self._get_unique_id(row, unique_index, row_idx)
            unique_prefix = f"[ID:{unique_id}] "
            
            row_content = []
            current_chunk = unique_prefix
            current_chunk_length = len(unique_prefix)
            
            for col_idx, cell_value in enumerate(row):
                # 获取列名（使用表头或列字母）
                if headers and col_idx < len(headers):
                    field = headers[col_idx]
                else:
                    field = openpyxl.utils.get_column_letter(col_idx + 1)
                
                # 处理合并单元格值
                if (row_idx, col_idx + 1) in merged_map:
                    main_row, main_col = merged_map[(row_idx, col_idx + 1)]
                    # 如果是合并单元格但不是主单元格，则使用主单元格的值
                    if (row_idx, col_idx + 1) != (main_row, main_col):
                        cell_value = self._get_merged_cell_value(sheet, main_row, main_col)
                
                # 处理特殊内容（链接、图片等）
                processed_value = self._process_cell_value(cell_value, sheet.cell(row=row_idx, column=col_idx + 1))
                
                # 清理特殊字符
                sanitized_value = self._sanitize_value(str(processed_value) if processed_value is not None else "")
                
                # 构建字段内容
                field_header = f"{field}: "
                field_header_length = len(field_header)
                item = f"{field_header}{sanitized_value}"
                item_length = len(item)
                
                # 检查是否超长需要分段
                if len(unique_prefix) + item_length > MAX_LEN - OVERHEAD_PADDING:
                    segments = self._segment_long_field(
                        unique_prefix, field, sanitized_value, 
                        MAX_LEN, field_header_length, OVERHEAD_PADDING
                    )
                    row_content.extend(segments)
                    continue
                
                # 检查当前块空间
                if current_chunk_length + item_length + 2 > MAX_LEN:  # +2 为分隔符预留
                    row_content.append(current_chunk.rstrip('; '))
                    current_chunk = unique_prefix + item + '; '
                    current_chunk_length = len(unique_prefix) + item_length + 2
                else:
                    current_chunk += item + '; '
                    current_chunk_length += item_length + 2
            
            # 添加剩余内容
            if current_chunk_length > len(unique_prefix):
                row_content.append(current_chunk.rstrip('; '))
            
            content.extend(row_content)
        
        return '\n'.join(content)
    
    def _get_merged_cell_value(self, sheet, row: int, col: int):
        """获取合并单元格主单元格的值"""
        try:
            # 对于只读模式，我们需要重新获取单元格值
            if hasattr(sheet, 'cell'):
                return sheet.cell(row=row, column=col).value
            return None
        except:
            return None
    
    def _get_headers(self, sheet) -> List[str]:
        """获取表头（第一行）"""
        headers = []
        try:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                headers = [str(cell) if cell is not None else f"Column{idx+1}" 
                          for idx, cell in enumerate(first_row)]
        except StopIteration:
            pass
        return headers
    
    def _get_unique_id(self, row, unique_index, row_idx) -> str:
        """获取行唯一标识"""
        if row and unique_index < len(row) and row[unique_index] is not None:
            return str(row[unique_index])
        return f"ROW{row_idx}"
    
    def _process_cell_value(self, value, cell=None) -> str:
        """处理特殊单元格内容"""
        # 处理超链接 - 检查cell是否有hyperlink属性
        if cell and hasattr(cell, 'hyperlink') and cell.hyperlink:
            return f"{value or ''} [Link:{cell.hyperlink.target or cell.hyperlink.location}]"
        
        # 处理公式生成的图片
        if isinstance(value, str) and value.startswith('=IMAGE('):
            return "[Image]"
        
        # 处理布尔值
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        
        return value
    
    def _segment_long_field(self, unique_prefix: str, field: str, value: str, 
                           max_len: int, field_header_length: int, padding: int) -> List[str]:
        """分段处理超长字段值"""
        segments = []
        segment_idx = 1
        max_segment_value_length = max_len - len(unique_prefix) - len(f"{field} [Part{segment_idx}]: ") - padding
        
        start = 0
        while start < len(value):
            segment_header = f"{field} [Part{segment_idx}]: "
            max_segment_value_length = max_len - len(unique_prefix) - len(segment_header) - padding
            
            if max_segment_value_length <= 0:
                max_segment_value_length = 1
                
            end = start + max_segment_value_length
            segment_value = value[start:end]
            
            segments.append(unique_prefix + segment_header + segment_value)
            
            start = end
            segment_idx += 1
        
        return segments
    
    def _sanitize_value(self, value: str) -> str:
        """清理值中的特殊字符"""
        replacements = {
            ';': '；',
            '\n': '↵',
            '\r': '',
            '\t': '    '
        }
        for orig, repl in replacements.items():
            value = value.replace(orig, repl)
        return value
    
    def _check_has_formulas(self) -> bool:
        """检查是否包含公式"""
        for sheet in self.workbook:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        return True
        return False
    
    def _check_has_hyperlinks(self) -> bool:
        """检查是否包含超链接"""
        for sheet in self.workbook:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        return True
        return False
    
    def _check_has_merged_cells(self) -> bool:
        """检查是否包含合并单元格"""
        for sheet in self.workbook:
            # 兼容只读模式
            if hasattr(sheet, 'merged_cells') and sheet.merged_cells.ranges:
                return True
            if hasattr(sheet, 'merged_cell_ranges') and sheet.merged_cell_ranges:
                return True
        return False